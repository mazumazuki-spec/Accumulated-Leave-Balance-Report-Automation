"""Phase 3: consolidate branch reports into one summary workbook."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import range_boundaries

from config_loader import AppConfig, BranchConfig, load_config
from excel_utils import auto_fit_columns, ensure_folder, parse_number, safe_load_workbook


def resolve_report_file(config: AppConfig, branch: BranchConfig) -> Path:
    """Return the branch report workbook created by Phase 2."""

    expected_file = (
        config.output_dir
        / config.to_date
        / "reports"
        / config.report_test_file_pattern.format(branch=branch.name, date=config.to_date)
    )
    if expected_file.exists():
        return expected_file

    raise FileNotFoundError(f"Report workbook not found: {expected_file}")


def read_branch_summary(config: AppConfig, branch: BranchConfig) -> dict[str, float | str]:
    """Read summary values from one branch report."""

    workbook = safe_load_workbook(resolve_report_file(config, branch), data_only=False)
    worksheet = workbook["Report_Test"] if "Report_Test" in workbook.sheetnames else workbook.active

    _, min_row, _, _ = range_boundaries(branch.report_range)

    summary = {
        "branch": branch.name,
        "total_employees": worksheet.cell(row=min_row, column=5).value,
        "above_from_date": worksheet.cell(row=min_row + 1, column=5).value,
        "above_current_date": worksheet.cell(row=min_row + 2, column=5).value,
    }

    from_total = 0.0
    current_total = 0.0

    for col in range(6, 10):
        value = parse_number(worksheet.cell(row=min_row, column=col).value)
        if value is not None:
            from_total += value

    for col in range(10, 14):
        value = parse_number(worksheet.cell(row=min_row, column=col).value)
        if value is not None:
            current_total += value

    summary["from_total"] = from_total
    summary["current_total"] = current_total

    workbook.close()
    return summary


def get_change_wording(config: AppConfig, from_total: float, current_total: float) -> tuple[str, str]:
    """Return change wording based on from-date and current-date totals."""

    wording = config.consolidated_summary

    if from_total > current_total:
        return wording.decrease_wording["amount"], wording.decrease_wording["percent"]

    if current_total > from_total:
        return wording.increase_wording["amount"], wording.increase_wording["percent"]

    return wording.equal_wording["amount"], wording.equal_wording["percent"]


def create_consolidated_summary(config: AppConfig) -> Path:
    """Create a consolidated workbook from all branch summaries."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Consolidated Summary"

    worksheet["A1"] = "Demo Consolidated Leave Balance Summary"
    worksheet["A2"] = "From Date"
    worksheet["B2"] = config.from_date
    worksheet["A3"] = "Current Date"
    worksheet["B3"] = config.to_date

    headers = [
        "Branch",
        "Total Employees",
        "Employees Above Threshold From Date",
        "Employees Above Threshold Current Date",
        "From Date Total",
        "Current Date Total",
    ]
    worksheet.append([])
    worksheet.append(headers)

    start_data_row = 6

    for index, branch in enumerate(config.branches, start=start_data_row):
        summary = read_branch_summary(config, branch)
        worksheet.cell(row=index, column=1).value = summary["branch"]
        worksheet.cell(row=index, column=2).value = summary["total_employees"]
        worksheet.cell(row=index, column=3).value = summary["above_from_date"]
        worksheet.cell(row=index, column=4).value = summary["above_current_date"]
        worksheet.cell(row=index, column=5).value = summary["from_total"]
        worksheet.cell(row=index, column=6).value = summary["current_total"]

    total_row = start_data_row + len(config.branches)
    worksheet.cell(row=total_row, column=1).value = "Grand Total"
    worksheet.cell(row=total_row, column=5).value = f"=SUM(E{start_data_row}:E{total_row - 1})"
    worksheet.cell(row=total_row, column=6).value = f"=SUM(F{start_data_row}:F{total_row - 1})"

    from_total = sum(
        parse_number(worksheet.cell(row=row, column=5).value) or 0
        for row in range(start_data_row, total_row)
    )
    current_total = sum(
        parse_number(worksheet.cell(row=row, column=6).value) or 0
        for row in range(start_data_row, total_row)
    )
    amount_wording, percent_wording = get_change_wording(config, from_total, current_total)

    worksheet["H5"] = "Change Amount Wording"
    worksheet["I5"] = amount_wording
    worksheet["H6"] = "Change Percent Wording"
    worksheet["I6"] = percent_wording

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    auto_fit_columns(worksheet)

    output_folder = ensure_folder(config.output_dir / config.to_date / "summary")
    output_file = output_folder / config.consolidated_summary.output_file.format(
        date=config.to_date,
    )

    workbook.save(output_file)
    workbook.close()

    print(f"Consolidated summary created: {output_file}")
    return output_file


def run_phase(config: AppConfig) -> Path:
    """Run Phase 3."""

    return create_consolidated_summary(config)


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments."""

    parser = argparse.ArgumentParser(description="Phase 3: consolidate summary.")
    parser.add_argument("--config", default="config/config.example.json")
    return parser.parse_args()


def main() -> None:
    """CLI entry point."""

    args = parse_args()
    config = load_config(Path(args.config))
    run_phase(config)


if __name__ == "__main__":
    main()
