"""Create sample input and template workbooks for the demo workflow."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config_loader import load_config
from excel_utils import auto_fit_columns, ensure_folder


def create_raw_workbook(file_path: Path) -> None:
    """Create a dummy raw workbook for demo processing."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Raw Data"

    headers = [
        "Employee ID",
        "Employee Name",
        "Department",
        "Position",
        "Leave Balance From Date",
        "Used Leave From Date",
        "Remaining From Date",
        "Leave Balance Current Date",
        "Used Leave Current Date",
        "Remaining Current Date",
        "Status",
    ]
    worksheet.append(headers)

    rows = [
        ["EMP_DEMO_001", "Demo Employee 1", "Operations", "Staff", 8, 2, 6, 10, 3, 7, "Active"],
        ["EMP_DEMO_002", "Demo Employee 2", "Operations", "Supervisor", 6, 1, 5, 4, 1, 3, "Active"],
        ["EMP_DEMO_003", "Demo Employee 3", "Support", "Staff", 1, 0, 1, 9, 2, 7, "Active"],
        ["EMP_DEMO_999", "Excluded Demo", "Support", "Staff", 12, 0, 12, 12, 0, 12, "Inactive"],
        ["EMP_DEMO_004", "Construction Demo", "Construction", "Temp", 6, 0, 6, 6, 0, 6, "Inactive"],
        ["Grand Total", "", "", "", 33, 3, 30, 41, 6, 35, ""],
    ]

    for row in rows:
        worksheet.append(row)

    auto_fit_columns(worksheet)
    ensure_folder(file_path.parent)
    workbook.save(file_path)
    workbook.close()


def create_template_workbook(file_path: Path, report_range: str) -> None:
    """Create a dummy report template workbook."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"

    worksheet["A1"] = "Demo Branch Report"
    worksheet["A1"].font = Font(bold=True, size=14)

    min_cell, max_cell = report_range.split(":")
    start_row = int("".join(filter(str.isdigit, min_cell)))
    start_col_letter = "".join(filter(str.isalpha, min_cell))

    worksheet[f"{start_col_letter}{start_row}"] = "Total Employees"
    worksheet[f"{start_col_letter}{start_row + 1}"] = "Above Threshold From Date"
    worksheet[f"{start_col_letter}{start_row + 2}"] = "Above Threshold Current Date"

    for row in range(start_row, start_row + 3):
        for col in range(4, 16):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.merge_cells(start_row=start_row, start_column=6, end_row=start_row, end_column=9)
    worksheet.cell(row=start_row, column=6).value = "From Date Total"
    worksheet.merge_cells(start_row=start_row, start_column=10, end_row=start_row, end_column=13)
    worksheet.cell(row=start_row, column=10).value = "Current Date Total"

    auto_fit_columns(worksheet)
    ensure_folder(file_path.parent)
    workbook.save(file_path)
    workbook.close()


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments."""

    parser = argparse.ArgumentParser(description="Create demo Excel files.")
    parser.add_argument(
        "--config",
        default="config/config.example.json",
        help="Path to config JSON file.",
    )
    return parser.parse_args()


def main() -> None:
    """Create sample raw files and template files."""

    args = parse_args()
    config = load_config(Path(args.config))

    ensure_folder(config.input_dir)
    ensure_folder(config.template_dir)

    for branch in config.branches:
        raw_file = config.input_dir / config.raw_file_pattern.format(
            branch=branch.name,
            date=config.to_date,
        )
        template_file = config.template_dir / f"{branch.name}_template.xlsx"

        create_raw_workbook(raw_file)
        create_template_workbook(template_file, branch.report_range)

        print(f"Created raw file: {raw_file}")
        print(f"Created template file: {template_file}")


if __name__ == "__main__":
    main()
