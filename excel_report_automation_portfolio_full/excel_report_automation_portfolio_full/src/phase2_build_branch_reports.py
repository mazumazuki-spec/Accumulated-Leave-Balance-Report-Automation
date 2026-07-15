"""Phase 2: build branch-level reports from cleaned files and templates."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.filters import CustomFilter, CustomFilters, FilterColumn

from config_loader import AppConfig, BranchConfig, load_config
from excel_utils import (
    apply_basic_table_format,
    auto_fit_columns,
    copy_range_with_style,
    ensure_folder,
    find_header_column,
    parse_number,
    safe_load_workbook,
)


def resolve_clean_file(config: AppConfig, branch: BranchConfig) -> Path:
    """Return the cleaned workbook for a branch."""

    expected_file = (
        config.output_dir
        / config.to_date
        / "cleaned"
        / config.clean_file_pattern.format(branch=branch.name, date=config.to_date)
    )
    if expected_file.exists():
        return expected_file

    raise FileNotFoundError(f"Cleaned workbook not found: {expected_file}")


def resolve_template_file(config: AppConfig, branch: BranchConfig) -> Path:
    """Return the report template workbook for a branch."""

    expected_file = config.template_dir / f"{branch.name}_template.xlsx"
    if expected_file.exists():
        return expected_file

    raise FileNotFoundError(f"Template workbook not found: {expected_file}")


def create_report_sheet(workbook):
    """Copy the cleaned sheet into a report test sheet."""

    if "Report_Test" in workbook.sheetnames:
        del workbook["Report_Test"]

    source_sheet = workbook["Cleaned"] if "Cleaned" in workbook.sheetnames else workbook.active
    report_sheet = workbook.copy_worksheet(source_sheet)
    report_sheet.title = "Report_Test"

    return report_sheet


def write_summary_formulas(worksheet, report_range: str, threshold_days: float) -> None:
    """Write generic summary formulas under the report range."""

    min_col, min_row, max_col, _ = range_boundaries(report_range)
    data_start_row = 2
    data_end_row = min_row - 1

    if data_end_row < data_start_row:
        return

    employee_id_col = find_header_column(
        worksheet,
        keywords=["employee id", "employee code", "staff id"],
        fallback_column=1,
    )
    employee_id_letter = get_column_letter(employee_id_col)

    from_balance_col = find_header_column(
        worksheet,
        keywords=["remaining from date", "balance from"],
        fallback_column=7,
    )
    current_balance_col = find_header_column(
        worksheet,
        keywords=["remaining current date", "balance current"],
        fallback_column=10,
    )

    from_balance_letter = get_column_letter(from_balance_col)
    current_balance_letter = get_column_letter(current_balance_col)

    total_row = min_row
    from_threshold_row = min_row + 1
    current_threshold_row = min_row + 2

    worksheet.cell(row=total_row, column=5).value = (
        f'=COUNTA(${employee_id_letter}${data_start_row}:'
        f'${employee_id_letter}${data_end_row})&" employees"'
    )
    worksheet.cell(row=from_threshold_row, column=5).value = (
        f'=COUNTIFS(${employee_id_letter}${data_start_row}:'
        f'${employee_id_letter}${data_end_row},"<>",'
        f'${from_balance_letter}${data_start_row}:'
        f'${from_balance_letter}${data_end_row},">{threshold_days:g}")&" employees"'
    )
    worksheet.cell(row=current_threshold_row, column=5).value = (
        f'=COUNTIFS(${employee_id_letter}${data_start_row}:'
        f'${employee_id_letter}${data_end_row},"<>",'
        f'${current_balance_letter}${data_start_row}:'
        f'${current_balance_letter}${data_end_row},">{threshold_days:g}")&" employees"'
    )

    for col in range(6, min(max_col, 13) + 1):
        col_letter = get_column_letter(col)
        worksheet.cell(row=total_row, column=col).value = (
            f"=SUM({col_letter}${data_start_row}:{col_letter}${data_end_row})"
        )


def apply_threshold_filter(worksheet, report_range: str, threshold_days: float) -> None:
    """Apply AutoFilter and hide rows not above threshold."""

    _, min_row, max_col, _ = range_boundaries(report_range)
    data_end_row = min_row - 1

    if data_end_row < 2:
        return

    current_balance_col = find_header_column(
        worksheet,
        keywords=["remaining current date", "balance current"],
        fallback_column=10,
    )
    current_balance_letter = get_column_letter(current_balance_col)

    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{data_end_row}"
    worksheet.auto_filter.filterColumn = []

    filter_column = FilterColumn(
        colId=current_balance_col - 1,
        customFilters=CustomFilters(
            customFilter=[
                CustomFilter(operator="greaterThan", val=str(threshold_days)),
            ],
        ),
    )
    worksheet.auto_filter.filterColumn.append(filter_column)

    for row in range(2, data_end_row + 1):
        value = worksheet.cell(row=row, column=current_balance_col).value
        numeric_value = parse_number(value)

        worksheet.row_dimensions[row].hidden = (
            numeric_value is None or numeric_value <= threshold_days
        )

    print(
        f"Applied filter on {current_balance_letter}: "
        f"greater than {threshold_days:g}"
    )


def update_condition_text(worksheet, threshold_days: float, from_date: str, to_date: str) -> None:
    """Update demo condition wording inside report labels."""

    replacements = {
        "{threshold_days}": f"{threshold_days:g}",
        "{from_date}": from_date,
        "{to_date}": to_date,
    }

    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                new_value = cell.value
                for old, new in replacements.items():
                    new_value = new_value.replace(old, new)
                cell.value = new_value


def build_branch_report(config: AppConfig, branch: BranchConfig) -> Path:
    """Build a report workbook for one branch."""

    clean_file = resolve_clean_file(config, branch)
    template_file = resolve_template_file(config, branch)

    clean_workbook = safe_load_workbook(clean_file)
    template_workbook = safe_load_workbook(template_file)

    template_sheet = template_workbook["Report"] if "Report" in template_workbook.sheetnames else template_workbook.active
    report_sheet = create_report_sheet(clean_workbook)

    copy_range_with_style(template_sheet, report_sheet, branch.report_range)
    write_summary_formulas(report_sheet, branch.report_range, config.summary_threshold_days)
    apply_threshold_filter(report_sheet, branch.report_range, config.summary_threshold_days)
    update_condition_text(
        report_sheet,
        threshold_days=config.summary_threshold_days,
        from_date=config.from_date,
        to_date=config.to_date,
    )

    _, min_row, max_col, _ = range_boundaries(branch.report_range)
    apply_basic_table_format(report_sheet, 1, max(min_row - 1, 1), 1, max_col)
    auto_fit_columns(report_sheet)

    output_folder = ensure_folder(config.output_dir / config.to_date / "reports")
    output_file = output_folder / config.report_test_file_pattern.format(
        branch=branch.name,
        date=config.to_date,
    )

    clean_workbook.save(output_file)
    clean_workbook.close()
    template_workbook.close()

    print(f"[{branch.name}] report created: {output_file}")
    return output_file


def run_phase(config: AppConfig) -> list[Path]:
    """Run Phase 2 for all branches."""

    return [build_branch_report(config, branch) for branch in config.branches]


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments."""

    parser = argparse.ArgumentParser(description="Phase 2: build branch reports.")
    parser.add_argument("--config", default="config/config.example.json")
    return parser.parse_args()


def main() -> None:
    """CLI entry point."""

    args = parse_args()
    config = load_config(Path(args.config))
    run_phase(config)


if __name__ == "__main__":
    main()
