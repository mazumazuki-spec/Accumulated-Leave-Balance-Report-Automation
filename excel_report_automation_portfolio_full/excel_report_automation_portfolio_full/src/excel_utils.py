"""Reusable Excel helper functions for the demo workflow."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet


def normalize_value(value: Any) -> str:
    """Convert an Excel value to clean comparable text."""

    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def normalize_employee_id(value: Any) -> str:
    """Normalize employee ID values from Excel."""

    text = normalize_value(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text


def parse_number(value: Any) -> float | None:
    """Parse a numeric value from Excel or text."""

    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if not text or text.startswith("="):
        return None

    match = re.search(r"-?\d+(\.\d+)?", text)
    if not match:
        return None

    return float(match.group(0))


def display_date(date_text: str, output_format: str = "%d-%b-%Y") -> str:
    """Format a YYYY-MM-DD date string for report display."""

    return datetime.strptime(date_text, "%Y-%m-%d").strftime(output_format)


def ensure_folder(path: Path) -> Path:
    """Create a folder if it does not exist."""

    path.mkdir(parents=True, exist_ok=True)
    return path


def get_first_visible_sheet(workbook) -> Worksheet:
    """Return the first visible sheet in a workbook."""

    for worksheet in workbook.worksheets:
        if worksheet.sheet_state == "visible":
            return worksheet
    return workbook.active


def copy_visible_sheet(workbook, target_title: str) -> Worksheet:
    """Copy the first visible sheet into a new sheet."""

    if target_title in workbook.sheetnames:
        del workbook[target_title]

    source_sheet = get_first_visible_sheet(workbook)
    target_sheet = workbook.copy_worksheet(source_sheet)
    target_sheet.title = target_title
    workbook.active = workbook.sheetnames.index(target_title)

    return target_sheet


def row_values_as_text(worksheet: Worksheet, row: int) -> list[str]:
    """Return normalized row values."""

    return [
        normalize_value(worksheet.cell(row=row, column=column).value)
        for column in range(1, worksheet.max_column + 1)
    ]


def row_contains_keyword(worksheet: Worksheet, row: int, keyword: str) -> bool:
    """Check whether any row value contains the keyword."""

    keyword_lower = keyword.lower()
    return any(keyword_lower in value.lower() for value in row_values_as_text(worksheet, row))


def row_contains_any_id(
    worksheet: Worksheet,
    row: int,
    employee_ids: Iterable[str],
) -> bool:
    """Check whether a row contains any exact employee ID."""

    id_set = {str(employee_id).strip() for employee_id in employee_ids}
    row_values = [normalize_employee_id(value) for value in row_values_as_text(worksheet, row)]
    return any(value in id_set for value in row_values)


def delete_rows_by_keywords(worksheet: Worksheet, keywords: Iterable[str]) -> int:
    """Delete rows containing configured keywords."""

    keywords = [keyword for keyword in keywords if keyword]
    deleted_count = 0

    for row in range(worksheet.max_row, 0, -1):
        if any(row_contains_keyword(worksheet, row, keyword) for keyword in keywords):
            worksheet.delete_rows(row)
            deleted_count += 1

    return deleted_count


def delete_rows_by_employee_ids(worksheet: Worksheet, employee_ids: Iterable[str]) -> int:
    """Delete rows containing configured employee IDs."""

    employee_ids = list(employee_ids)
    if not employee_ids:
        return 0

    deleted_count = 0

    for row in range(worksheet.max_row, 0, -1):
        if row_contains_any_id(worksheet, row, employee_ids):
            worksheet.delete_rows(row)
            deleted_count += 1

    return deleted_count


def delete_columns_by_letters(worksheet: Worksheet, column_letters: Iterable[str]) -> None:
    """Delete configured Excel columns by letter."""

    column_indexes = []
    for column_letter in column_letters:
        try:
            column_indexes.append(worksheet[f"{column_letter}1"].column)
        except ValueError:
            continue

    for column_index in sorted(set(column_indexes), reverse=True):
        if column_index <= worksheet.max_column:
            worksheet.delete_cols(column_index, 1)


def find_header_column(
    worksheet: Worksheet,
    keywords: Iterable[str],
    max_scan_rows: int = 30,
    fallback_column: int = 2,
) -> int:
    """Find a column by scanning header rows for keywords."""

    keyword_list = [keyword.lower() for keyword in keywords]

    for row in range(1, min(worksheet.max_row, max_scan_rows) + 1):
        for column in range(1, worksheet.max_column + 1):
            text = normalize_value(worksheet.cell(row=row, column=column).value).lower()
            if any(keyword in text for keyword in keyword_list):
                return column

    return fallback_column


def auto_fit_columns(
    worksheet: Worksheet,
    min_width: int = 8,
    max_width: int = 45,
    skip_formula: bool = True,
) -> None:
    """Auto fit worksheet columns based on text length."""

    for column in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column)
        max_length = 0

        for row in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row, column=column).value
            if value is None:
                continue

            text = str(value)
            if skip_formula and text.startswith("="):
                continue

            longest_line = max(len(line) for line in text.split("\n"))
            max_length = max(max_length, longest_line)

        worksheet.column_dimensions[column_letter].width = max(
            min_width,
            min(max_length + 2, max_width),
        )


def apply_basic_table_format(
    worksheet: Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    """Apply a simple table format to a worksheet range."""

    if end_row < start_row:
        return

    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    highlight_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    thin_side = Side(style="thin", color="000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in range(start_row, end_row + 1):
        for column in range(start_col, end_col + 1):
            worksheet.cell(row=row, column=column).border = border

    for column in range(start_col, end_col + 1):
        worksheet.cell(row=start_row, column=column).fill = header_fill

    if end_col >= start_col:
        worksheet.cell(row=start_row, column=end_col).fill = highlight_fill


def copy_range_with_style(source_ws: Worksheet, target_ws: Worksheet, source_range: str) -> None:
    """Copy values and styles from a source range to the same range in a target sheet."""

    min_col, min_row, max_col, max_row = range_boundaries(source_range)

    for merged_range in list(target_ws.merged_cells.ranges):
        r_min_col, r_min_row, r_max_col, r_max_row = merged_range.bounds
        intersects = not (
            r_max_row < min_row
            or r_min_row > max_row
            or r_max_col < min_col
            or r_min_col > max_col
        )
        if intersects:
            target_ws.unmerge_cells(str(merged_range))

    for row in range(min_row, max_row + 1):
        target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height
        for column in range(min_col, max_col + 1):
            source_cell = source_ws.cell(row=row, column=column)
            target_cell = target_ws.cell(row=row, column=column)
            target_cell.value = source_cell.value

            if source_cell.has_style:
                from copy import copy

                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)

    for column in range(min_col, max_col + 1):
        column_letter = get_column_letter(column)
        target_ws.column_dimensions[column_letter].width = source_ws.column_dimensions[
            column_letter
        ].width

    for merged_range in list(source_ws.merged_cells.ranges):
        r_min_col, r_min_row, r_max_col, r_max_row = merged_range.bounds
        inside = (
            r_min_row >= min_row
            and r_max_row <= max_row
            and r_min_col >= min_col
            and r_max_col <= max_col
        )
        if inside:
            target_ws.merge_cells(str(merged_range))


def write_merged_safe(worksheet: Worksheet, cell_address: str, value: Any) -> None:
    """Write a value into a cell safely when merged cells may be present."""

    cell = worksheet[cell_address]

    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            top_left.value = value
            return

    cell.value = value


def safe_load_workbook(file_path: Path):
    """Load an Excel workbook with a clear error message."""

    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    return load_workbook(file_path)
