"""Phase 1: download or read raw Excel files and clean branch data."""

from __future__ import annotations

import argparse
from pathlib import Path

import requests

from config_loader import AppConfig, BranchConfig, load_config
from excel_utils import (
    auto_fit_columns,
    copy_visible_sheet,
    delete_columns_by_letters,
    delete_rows_by_employee_ids,
    delete_rows_by_keywords,
    ensure_folder,
    find_header_column,
    normalize_employee_id,
    safe_load_workbook,
)
from openpyxl.styles import Alignment, PatternFill


def build_download_url(config: AppConfig, branch: BranchConfig) -> str:
    """Build a configurable demo API URL for one branch."""

    if not config.api_base_url:
        raise ValueError("api_base_url is empty. Use local files or set a demo URL in config.")

    return (
        f"{config.api_base_url}"
        f"?branch_id={branch.branch_id}"
        f"&from_date={config.from_date}"
        f"&to_date={config.to_date}"
    )


def download_raw_file(config: AppConfig, branch: BranchConfig, output_file: Path) -> None:
    """Download one branch raw file from a configurable API endpoint."""

    url = build_download_url(config, branch)
    response = requests.get(url, timeout=60)
    response.raise_for_status()

    ensure_folder(output_file.parent)
    output_file.write_bytes(response.content)


def resolve_raw_file(config: AppConfig, branch: BranchConfig) -> Path:
    """Resolve the raw input file for a branch."""

    expected_file = config.input_dir / config.raw_file_pattern.format(
        branch=branch.name,
        date=config.to_date,
    )
    if expected_file.exists():
        return expected_file

    matches = list(config.input_dir.glob(f"*{branch.name}*.xlsx"))
    if matches:
        return matches[0]

    raise FileNotFoundError(f"No raw input file found for {branch.name}: {expected_file}")


def add_employee_remarks(worksheet, remark_employee_ids: dict[str, str]) -> int:
    """Insert a remark column and fill remark text for configured employee IDs."""

    if not remark_employee_ids:
        return 0

    employee_id_col = find_header_column(
        worksheet,
        keywords=["employee id", "employee code", "staff id"],
        fallback_column=1,
    )
    position_col = find_header_column(
        worksheet,
        keywords=["position", "job title"],
        fallback_column=4,
    )

    remark_col = position_col + 1
    worksheet.insert_cols(remark_col)

    if employee_id_col >= remark_col:
        employee_id_col += 1

    header_cell = worksheet.cell(row=1, column=remark_col)
    header_cell.value = "Remark"
    header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    updated_count = 0

    for row in range(2, worksheet.max_row + 1):
        employee_id = normalize_employee_id(worksheet.cell(row=row, column=employee_id_col).value)
        if employee_id in remark_employee_ids:
            worksheet.cell(row=row, column=remark_col).value = remark_employee_ids[employee_id]
            for column in range(1, remark_col + 1):
                worksheet.cell(row=row, column=column).fill = fill
            updated_count += 1

    return updated_count


def clean_branch_file(config: AppConfig, branch: BranchConfig, raw_file: Path) -> Path:
    """Clean one branch workbook and save a cleaned workbook."""

    workbook = safe_load_workbook(raw_file)
    worksheet = copy_visible_sheet(workbook, "Cleaned")

    deleted_keyword_rows = delete_rows_by_keywords(worksheet, config.delete_keywords)
    deleted_employee_rows = delete_rows_by_employee_ids(
        worksheet,
        branch.exclude_employee_ids,
    )

    if config.delete_columns:
        delete_columns_by_letters(worksheet, config.delete_columns)

    remarks_added = add_employee_remarks(worksheet, branch.remark_employee_ids)
    auto_fit_columns(worksheet)

    output_folder = ensure_folder(config.output_dir / config.to_date / "cleaned")
    output_file = output_folder / config.clean_file_pattern.format(
        branch=branch.name,
        date=config.to_date,
    )

    workbook.save(output_file)
    workbook.close()

    print(
        f"[{branch.name}] cleaned: {output_file} | "
        f"keyword rows deleted={deleted_keyword_rows}, "
        f"employee rows deleted={deleted_employee_rows}, "
        f"remarks added={remarks_added}"
    )

    return output_file


def run_phase(config: AppConfig, use_download: bool = False) -> list[Path]:
    """Run Phase 1 for all configured branches."""

    output_files: list[Path] = []
    ensure_folder(config.input_dir)

    for branch in config.branches:
        if use_download:
            raw_file = config.input_dir / config.raw_file_pattern.format(
                branch=branch.name,
                date=config.to_date,
            )
            download_raw_file(config, branch, raw_file)
        else:
            raw_file = resolve_raw_file(config, branch)

        output_files.append(clean_branch_file(config, branch, raw_file))

    return output_files


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments."""

    parser = argparse.ArgumentParser(description="Phase 1: clean raw Excel files.")
    parser.add_argument("--config", default="config/config.example.json")
    parser.add_argument("--download", action="store_true")
    return parser.parse_args()


def main() -> None:
    """CLI entry point."""

    args = parse_args()
    config = load_config(Path(args.config))
    run_phase(config, use_download=args.download)


if __name__ == "__main__":
    main()
